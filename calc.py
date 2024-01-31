from pals_data import Pal


XLSX_NAME = '帕鲁全配种计算.xlsx'

breed = {}  # 配种合成表
breed_detail = {}   # 配种合成表详细
breed_target = {}   # 配种合成表目标

def add_breed_detail(A, B, C):
    if A not in breed_detail:
        breed_detail[A] = {}
        breed_target[A] = []
    breed_detail[A][B] = C
    if C not in breed_target[A]:
        breed_target[A].append(C)
    
def include_xlsx():
    """从xlsx中读取数据"""
    # 加载帕鲁数据
    Pal.init()
    # 读配种表
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_NAME)
    # 打开第1个表 - 帕鲁配种表
    sheet = wb.worksheets[0]
    # 读取数据
    ## 表格从C3开始
    for row in sheet.iter_rows(min_row=3, min_col=3):
        for cell in row:
            if cell.value is None:
                continue
            column = cell.column
            A_name = sheet.cell(row=2, column=column).value
            B_name = sheet.cell(row=cell.row, column=2).value
            C_name = cell.value
            A = Pal.get_by_name(A_name)
            B = Pal.get_by_name(B_name)
            C = Pal.get_by_name(C_name)
            breed[(A, B)] = C
            add_breed_detail(A, B, C)
            

############################ COMMAND ############################
cmds = {}
def command(cmd : str, help : str):
    def decorator(func):
        cmds[cmd] = (func, help)
        return func
    return decorator

@command('count_end', '[count_end A] - 计算A的配种方案')
def calc_breed_count(A_no : str):
    """计算有多少种方案能合成该帕鲁"""
    name = Pal.get(A_no)
    ans = []
    for k , v in breed.items():
        if v == name and (k[1], k[0]) not in ans:
            ans.append(k)
    for a, b in ans:
        print(f"{a} + {b} = {name}")
    print(f"共有{len(ans)}种配种方案能合成{name}")

@command('road', '[road A B] - 计算从A到B的配种方案')
def calc_shortest_breed_road(A_no, B_no):
    """计算从A到B的最短合成路径"""
    A = Pal.get(A_no)
    B = Pal.get(B_no)
    print(f"计算从{A}到{B}的最短合成路径")
    import queue
    ret = []
    min_depth = 9999
    # 使用广度优先搜索
    q = queue.Queue()
    q.put(([A]))
    last_len = 0
    while not q.empty():
        road = q.get()
        if len(road) != last_len:
            print(f"开始计算{len(road)}步的路径，当前路径数：{q.qsize()}")
            last_len = len(road)
        if len(road) > 3:
            print("超过3步，视为找不到合成路径。")
            return
        current = road[-1]
        for target in breed_target[current]:
            if target == B and len(road) <= min_depth:
                # 找到了
                ret.append(road + [target])  
                min_depth = len(road)
            elif target in road:
                # 往回了，跳过
                continue
            else:
                if len(ret) == 0:
                    # 如果还没找到路径那就继续找，找到了就不再增加新的
                    new_road = road.copy()
                    new_road.append(target)
                    q.put(new_road)
    for x in ret:
        print(" -> ".join(str(i) for i in x))
    print(f"共有{len(ret)}种配种方案能从{A}合成{B}")

@command("calc_road_count", "[calc_road_count A] - 计算A能配出多少种帕鲁")
def calc_road_count(A_no):
    A = Pal.get(A_no)
    once = breed_target[A]
    twice = []
    all = []
    for name in once:
        for v in breed_target[name]:
            if v not in twice:
                twice.append(v)
    for x in once:
        if x not in all:
            all.append(x)
    for x in twice:
        if x not in all:
            all.append(x)
    print(f"{A}能1级配出{len(once)}种帕鲁，2级配出{len(twice)}种帕鲁，共{len(all)}种帕鲁。")

@command("sort_road_count", "[sort_road_count (max_level)] - 按照可配出数量排序")
def sort_road_count(max_level = 999):
    """按照可配出数量排序"""   
    # 1级涉及到3个帕鲁，A+B=C
    pals_1 = {}
    for a, a_detail in breed_detail.items():
        if a.min_level > int(max_level):
            continue
        for b, b_detail in breed_detail.items():
            if b.min_level > int(max_level):
                continue
            c = breed_detail[a][b]
            # a+b=c
            if a not in pals_1:
                pals_1[a] = []
            if b not in pals_1:
                pals_1[b] = []
            if c not in pals_1[a]:           
                pals_1[a].append(c)
            if c not in pals_1[b]:
                pals_1[b].append(c)
        
    # 2级涉及到5个帕鲁，A+B=D, C+D=E
    pals_2 = {}
    for a, a_detail in breed_detail.items():
        if a.min_level > int(max_level):
            continue
        for b, b_detail in breed_detail.items():
            if b.min_level > int(max_level):
                continue
            for c, c_detail in breed_detail.items():
                if c.min_level > int(max_level):
                    continue
                # 枚举3个帕鲁
                d = breed_detail[a][b]
                e = breed_detail[c][d]
                # 相当于a+b+c=e
                if a not in pals_2:
                    pals_2[a] = []
                if b not in pals_2:
                    pals_2[b] = []
                if c not in pals_2:
                    pals_2[c] = []
                if e not in pals_2[a]:
                    pals_2[a].append(e)
                if e not in pals_2[b]:
                    pals_2[b].append(e)
                if e not in pals_2[c]:
                    pals_2[c].append(e)
    print("1级数优先排序")
    count = []
    for pal, pal_1 in pals_1.items():
        count.append((pal, len(pal_1)))
    count.sort(key=lambda x: x[1], reverse=True)
    for i in range(20):
        if i >= len(count):
            break
        x = count[i][0]
        print(f"{i+1} : {x} - {len(pals_1[x])} - {len(pals_2[x])}")
    print("2级数优先排序")
    count = []
    for pal, pal_2 in pals_2.items():
        count.append((pal, len(pal_2)))
    count.sort(key=lambda x: x[1], reverse=True)
    for i in range(20):
        if i >= len(count):
            break
        x = count[i][0]
        print(f"{i+1} : {x} - {len(pals_1[x])} - {len(pals_2[x])}")
        

@command("help", "显示帮助")
def help():
    """显示帮助"""
    for cmd in cmds:
        print(cmds[cmd][1])


if __name__ == '__main__':
    include_xlsx()
    while True:
        try:
            cmd = input('请输入命令（帮助请输入help）：')
            args = cmd.split()
            for i in range(len(args)):
                args[i] = args[i].strip()
            if args[0] in cmds:
                cmds[args[0]][0](*args[1:])
            elif args[0] == "exit":
                break
            else:
                print('未知命令')
                help()
        except Exception as e:
            # 显示行、错误
            import traceback
            traceback.print_exc()
    
